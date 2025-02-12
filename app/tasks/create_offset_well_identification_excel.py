from tasks.task import Task
from tasks.task_enum import TASKS
from helpers import task_logger
from traceback import format_exc
import time
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
import os
from helpers import auto_adjust_column_widths
from services import AnalysisService, WellService, WellGroupService

class CreateOffsetWellIdentificationExcel(Task):
    
    def execute(self):
        task = TASKS.CREATE_OFFSET_WELL_IDENTIFICATION_EXCEL.value
        logger = task_logger(task, self.context.logs_path)
        try:
            analysis_service = AnalysisService(db_path=self.context.db_path)
            well_service = WellService(db_path=self.context.db_path)
            well_group_service = WellGroupService(db_path=self.context.db_path) 
            analyses = list(analysis_service.get_all_excluding_target_wells())

            df = pd.DataFrame(columns=['API_UWI', 
                                        'WellName', 
                                        'Co-dev',
                                        'Child',
                                        'Pct cum oil greater than 7.5',
                                        'In/Out',
                                        'Remarks',
                                        'Cum Oil',
                                        'Cum Oil bbl per ft',
                                        'Pct of Group Cum Oil bbl per ft',
                                        'ENVOperator', 
                                        'ENVInterval', 
                                        'FirstProdDate', 
                                        'TVD_FT', 
                                        'LateralLength_FT',
                                        'PerfInterval_FT', 
                                        'Effective_Perf_Interval'
                                        'ProppantIntensity_LBSPerFT', 
                                        'BH_Lat', 
                                        'BH_Long', 
                                        'RKB_Elev', 
                                        'TVD_SS',
                                        'Average-Lateral-Spacing-at-BHL',
                                        'Bound-Half-Bound',
                                        'Adjacent-Child',
                                        'Parents',
                                        'Parent-1',
                                        'Parent-1-First-Production-Date',
                                        'Parent_1-Delta-First-Production-Months',
                                        'Parent_1-Landing-Zone',
                                        'Parent-2',
                                        'Parent-2-First-Production-Date',
                                        'Parent_2-Delta-First-Production-Months',
                                        'Parent_2-Landing-Zone',
                                        'Adjacent-2-West',
                                        'Adjacent-2-Distance-West',
                                        'Adjacent-2-Hypotenuse-Distance-West',
                                        'Adjacent-1-East',
                                        'Adjacent-1-Distance-East',
                                        'Adjacent-1-Hypotenuse-Distance-East',
                                        'Group-ID', 
                                        'Group-Lateral-Spacing-at-BHL',
                                        'Group-Hypotenuse-Spacing-at-BHL'
                                        ], dtype="object")

            for analysis in analyses:
                well = well_service.get_by_name(analysis.name)
                row = {
                    'API_UWI': analysis.api,
                    'WellName': analysis.name,
                    'Co-dev': analysis.codevelopment,
                    'Child': analysis.child,
                    'Pct cum oil greater than 7.5': analysis.pct_of_group_cumoil_bblperft_greater_than,
                    'In/Out': None,
                    'Remarks': "",
                    'Cum Oil': well._cumlative_oil,
                    'Cum Oil bbl per ft': analysis.cumoil_bblperft,
                    'Pct of Group Cum Oil bbl per ft': analysis.pct_of_group_cumoil_bblperft,
                    'ENVOperator': well.operator,
                    'ENVInterval': well.interval,
                    'FirstProdDate': analysis.first_production_date,
                    'TVD_FT': well.total_vertical_depth,
                    'LateralLength_FT': well.lateral_length,
                    'PerfInterval_FT': well.perf_interval,
                    'Effective_Perf_Interval': well.perf_interval,
                    'ProppantIntensity_LBSPerFT': well.proppant_intensity,
                    'BH_Lat': well.bottom_hole_latitude,
                    'BH_Long': well.bottom_hole_longitude,
                    'RKB_Elev': well.kelly_bushing_elevation,
                    'TVD_SS': analysis.subsurface_depth,
                    'Average-Lateral-Spacing-at-BHL': analysis.average_horizontal_spacing,
                    'Bound-Half-Bound': analysis.bound,
                    'Adjacent-Child': analysis.adjacent_child,
                    'Parents': analysis.parents,
                    'Parent-1': analysis.parent_1,
                    'Parent-1-First-Production-Date': analysis.parent_1_first_production_date,
                    'Parent-1-Delta-First-Production-Months': analysis.parent_1_delta_first_production_months,
                    'Parent_1-Landing-Zone': analysis.parent_1_interval ,
                    'Parent-2': analysis.parent_2,
                    'Parent-2-First-Production-Date': analysis.parent_2_first_production_date,
                    'Parent_2-Delta-First-Production-Months': analysis.parent_2_delta_first_production_months,
                    'Parent_2-Landing-Zone': analysis.parent_2_interval,
                    'Adjacent-2-West': analysis.adjacent_2,
                    'Adjacent-2-Distance-West': analysis.distance_2,
                    'Adjacent-2-Hypotenuse-Distance-West': analysis.hypotenuse_2,
                    'Adjacent-1-East': analysis.adjacent_1,
                    'Adjacent-1-Distance-East': analysis.distance_1,
                    'Adjacent-1-Hypotenuse-Distance-East': analysis.hypotenuse_1,
                    'Group-ID': analysis.group_id,
                    'Group-Lateral-Spacing-at-BHL': analysis.group_average_horizontal_spacing,
                    'Group-Hypotenuse-Spacing-at-BHL': analysis.group_average_hypotenuse_spacing
                }

                row_df = pd.DataFrame([row])
                if not df.empty:
                    df = pd.concat([df, row_df], ignore_index=True)
                else:
                    df = row_df

            excel_file = os.path.join(self.context.project_path, f"{self.context.project}-offset-well-identification-{self.context.version}.xlsx")

            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)

            workbook = openpyxl.load_workbook(excel_file)
            worksheet = workbook.active

            # Freeze WellName column
            well_name_col_index = df.columns.get_loc('WellName') + 1
            freeze_col_letter = get_column_letter(well_name_col_index + 1)
            worksheet.freeze_panes = f'{freeze_col_letter}2'
            
            # Adjust the column widths
            auto_adjust_column_widths(worksheet)

            index = df.columns.get_loc('WellName') + 1
            column_letter = get_column_letter(index)
            if column_letter in worksheet.column_dimensions:
                column_width = worksheet.column_dimensions[column_letter].width
                if column_width is None:  # None means default width
                    column_width = worksheet.sheet_format.defaultColWidth
            else:
                column_width = worksheet.sheet_format.defaultColWidth 
            worksheet.column_dimensions[get_column_letter(df.columns.get_loc('Remarks')+1)].width = column_width

            last_row = worksheet.max_row
            last_column = worksheet.max_column
            last_column_letter = get_column_letter(last_column)
            worksheet.auto_filter.ref = f"A1:{last_column_letter}{last_row}"

            # Center align all cells
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # Add Work Groups worksheet
            work_group_worksheet = workbook.create_sheet(title="Work Groups", index=1)
            work_group_worksheet['A1'] = "Work Group"
            work_group_worksheet['B1'] = "Avg Cum Oil bbl per ft"
            work_groups = well_group_service.get_all()
            for i, work_group in enumerate(work_groups, start=2):
                work_group_worksheet[f'A{i}'] = work_group.name
                work_group_worksheet[f'B{i}'] = work_group.avg_cumoil_per_ft

            attempt_limit = 5
            attempt = 0

            while attempt < attempt_limit:
                try:
                    workbook.save(excel_file)
                    logger.info("File saved successfully.")
                    break  # Exit the loop if save was successful
                except PermissionError as e:
                    logger.warning(f"Attempt {attempt + 1} of {attempt_limit} failed: {e}")
                    time.sleep(2)  # Wait for 2 seconds before retrying
                    attempt += 1

            if attempt == attempt_limit:
                logger.error("Failed to save the file after several attempts.")
            logger.info(f"{task}: {self.context.logs_path}")
        except Exception as e:
            logger.error(f"Error {task} workflow task: {e}\n{format_exc()}")
            raise ValueError(f"Error {task} workflow task: {e}")
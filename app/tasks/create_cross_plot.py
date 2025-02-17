from tasks.task import Task
from tasks.task_enum import TASKS
from helpers import task_logger, HandlerWithText, calculate_angle
from services import (AnalysisService, 
                      TargetWellInformationService, 
                      WellService,
                      XYZDistanceService,
                      StratigraphicService,
                      TexasLandSurveySystemService,
                      NewMexicoLandSurveySystemService)

from traceback import format_exc
import matplotlib.pyplot as plt
import matplotlib.lines as mlines
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
import numpy as np
import mplcursors
import random
import colorsys

BACKGROUND_COLOR = "white"

class CreateCrossPlot(Task):

    def auto_adjust_column_widths(self, worksheet, headers):
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value in headers:
                        header_length = len(cell.value)
                        header_name = cell.value
                        continue
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            if max_length < header_length:
                adjusted_width = 12
            elif max_length > 25:
                adjusted_width = max_length
            else:
                adjusted_width = 25
            worksheet.column_dimensions[column].width = adjusted_width

    def execute(self):
        task = TASKS.CREATE_CROSS_PLOT.value
        logger = task_logger(task, self.context.logs_path)
        try:
            plot_name = "cross-plot"

            fontsize = 8

            analysis_service = AnalysisService(db_path=self.context.db_path)
            target_well_information_service = TargetWellInformationService(self.context.db_path)
            well_service = WellService(self.context.db_path)
            xyz_distance_service = XYZDistanceService(self.context.db_path)
            stratigraphic_service = StratigraphicService(self.context.db_path)
            texas_land_survey_service = TexasLandSurveySystemService(self.context._texas_land_survey_system_database_path)
            new_mexico_land_survey_service = NewMexicoLandSurveySystemService(self.context._new_mexico_land_survey_system_database_path)

            target_well = target_well_information_service.get_first_row()
            
            y_min = -9000
            y_max = -7000
            x_min = -30000
            x_max = 25000

            fig, ax = plt.subplots(figsize=(15, 8), facecolor=BACKGROUND_COLOR)
            ax.set_facecolor(BACKGROUND_COLOR)
            ax.grid(True, which='both', color=BACKGROUND_COLOR, linestyle='--', linewidth=0.5)
            
            title = f"{self.context.project.capitalize()} Cross Plot {self.context.version}"
            plt.title(title, fontsize=12, fontweight='bold')

            # Find X and Y limits
            target_wells = target_well_information_service.get_all()
            for target_well in target_wells:
                # if target_well.bhl_tvd_ss_ft < y_min:
                # y_min = round(target_well.bhl_tvd_ss_ft, -3) - 1500
                # if target_well.bhl_tvd_ss_ft > y_max:
                y_max = round(target_well.bhl_tvd_ss_ft, -3) + 1000

            ax.set_xlim(x_min, x_max)
            ax.set_ylim(y_min, y_max)

            # plt.yticks(np.arange(y_min, y_max + 1, 500))

            depths = np.arange(y_min, y_max, 100)
            for depth in depths:
                ax.axhline(depth, color='lightgrey', linewidth=0.5, linestyle='--', alpha=0.90, zorder=0)

            # plt.xticks(np.arange(x_min, x_max + 1, 5000))

            widths = np.arange(x_min, x_max, 1000)
            for width in widths:
                ax.axvline(width, color='lightgrey', linewidth=0.5, linestyle='--', alpha=0.90, zorder=0)  

            # Set axis labels
            if target_well.state == "TX":
                plss = texas_land_survey_service.get_by_county_abstract(target_well.county, target_well.tx_abstract_southwest_corner)           
                x_axis_label = f"Bottom hole spacing from west line {target_well.state}/{target_well.county}/{target_well.tx_abstract_southwest_corner}/{plss.block}/{int(float(plss.section))}) (500 ft/int.)"
            elif target_well.state == "NM":
                x_axis_label = f"Bottom hole spacing from west line {target_well.state}/{target_well.county}/{target_well.nw_township_southwest_corner}/{target_well.nm_range_southwest_corner}/{int(float(target_well.nm_tx_section_southwest_corner))}) (500 ft/int.)"
            ax.set_xlabel(x_axis_label)
            ax.set_ylabel(f"Depth below mean sea level (1000 ft/int.)")

            # Plot west section line
            specific_x = 0
            ax.axvline(specific_x, color='blue', linewidth=0.5, linestyle='--', alpha=0.75)
            label_y_position = y_min
            if target_well.state == "TX":
                ax.text(specific_x, label_y_position, x_axis_label, color='black', fontsize=fontsize, ha='center', va='bottom')
            elif target_well.state == "NM":
                ax.text(specific_x, label_y_position, x_axis_label, color='black', fontsize=fontsize, ha='center', va='bottom')
            
            xyz_distances = xyz_distance_service.get_by_simulated_well()
            
            colors = {}
            intervals = set()
            scatters = []
            well_name_list = []

            for i, xyz_distance in enumerate(xyz_distances, start=1):
                analysis = analysis_service.get_by_api(xyz_distance.target_api)

                # Skip if the depth is not within the plot range
                if analysis.subsurface_depth < y_min or analysis.subsurface_depth > y_max:
                    continue

                if "11-111" in xyz_distance.target_api:
                    color = "green"
                    alpha = 0.90
                    zorder = 5
                    marker = 's'  # 's' for square
                    size = 100
                else:
                    alpha = 0.70
                    zorder = 3
                    marker = 'o'  # 'o' for circle
                    size = 100

                    stratigraphic = stratigraphic_service.get_by_union_code(analysis.interval)
                    if stratigraphic:
                        color = stratigraphic.color
                    else:
                        color = "black"

                scatter = ax.scatter(xyz_distance.end_x, analysis.subsurface_depth, marker=marker, facecolor=color, s=size)
                ax.annotate(str(i), (xyz_distance.end_x, analysis.subsurface_depth), color='black', fontsize=fontsize, ha='center', va='center', weight='bold', zorder=6)

                # ax.scatter(xyz_distance.end_x, analysis.subsurface_depth, marker=marker, facecolor=color, s=size)
                scatters.append(scatter)
                well_name_list.append(xyz_distance.target_api)  

                if analysis.interval:
                    intervals.add(analysis.interval)
                    colors[analysis.interval] = color

            cursor = mplcursors.cursor(scatters, hover=True)

            @cursor.connect("add")
            def on_add(sel):
                # Fetch the index of the hovered scatter point and display its corresponding well name
                index = sel.index
                sel.annotation.set_text(well_name_list[index])
                sel.annotation.get_bbox_patch().set(fc="yellow", alpha=0.7)  # Optional: Customize the tooltip box

            #  Create custom legend entries
            circles = []
            interval_list = []
            for interval, color in colors.items():
                interval_list.append(interval)
            stratigraphics = stratigraphic_service.get_by_union_code_list(interval_list)

            circles.append(mlines.Line2D([], [], color='green', marker='s', markersize=10, label='Target Well', linestyle='None'))
            for stratigraphic in stratigraphics:
                circles.append(mlines.Line2D([], [], color=stratigraphic.color, marker='o', markersize=10, label=stratigraphic.union_code, linestyle='None')) 

            # Add the custom legend to the plot
            plt.legend(handles=circles, loc='center right', bbox_to_anchor=(1.15, 0.5), fontsize=12)

            # Save the plot
            output_file = os.path.join(self.context.project_path, f"{self.context.project}-{plot_name}-{self.context.version}.png")
            plt.savefig(output_file, dpi=300, bbox_inches='tight', facecolor=BACKGROUND_COLOR)
            plt.close(fig)

            # Save to excel
            cross_plot_workbook = Workbook()

            # Create Child Well Summary Worksheet
            plot_worksheet = cross_plot_workbook.active
            plot_worksheet.title = "Cross Plot"

            image = Image(output_file)
            image.width = image.width * 0.35 # Scale the width to 50%
            image.height = image.height * 0.35  # Scale the height to 50%
            plot_worksheet.add_image(image, f"A{plot_worksheet.max_row + 0}")

            # Save plot data to Excel
            data_worksheet = cross_plot_workbook.create_sheet("Data")            
            headers = [
                'Ref #',
                'Well API',
                'Well Name',
                'Well Landing Zone',
                'Feet from West Line',
                'Subsurface Depth'
            ]
            data_worksheet.append(headers)

            target_well_rows  = [int]

            for i, xyz_distance in enumerate(xyz_distances, start=1):
                analysis = analysis_service.get_by_api(xyz_distance.target_api)
                if analysis.subsurface_depth < y_min or analysis.subsurface_depth > y_max:
                    continue
                if "11-111" in xyz_distance.target_api:
                    target_well_rows.append(i)
                    target_well = target_well_information_service.get_by_name(analysis.name)
                    if target_well:
                        api = target_well.api if target_well.api else "Target Well"
                else:
                    api = analysis.api
                row = [
                    i,
                    api,
                    analysis.name,
                    analysis.interval,
                    xyz_distance.end_x,
                    analysis.subsurface_depth
                ]
                data_worksheet.append(row)

            # Adjust the column widths
            self.auto_adjust_column_widths(data_worksheet, headers)

            # Adjust background color for target well rows
            for target_well_row in target_well_rows:
                for row in data_worksheet.iter_rows():
                    if row[0].internal_value == target_well_row:
                        row_number = row[0].row
                        cell = data_worksheet[f"A{row_number}"]
                        cell.fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")

                    
            # Center align all cellschild_well_summary_worksheet
            for row in data_worksheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(left=Side(border_style='thin', color='FF000000'),
                                         right=Side(border_style='thin', color='FF000000'),
                                         top=Side(border_style='thin', color='FF000000'),
                                         bottom=Side(border_style='thin', color='FF000000'))

            # Set header backgroup color to light blue
            for cell in data_worksheet[1]:
                cell.fill = PatternFill(start_color="DDEBF7", end_color="00B0F0", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Save to excel
            excel_file = os.path.join(self.context.project_path, f"{self.context.project}-cross-plot-{self.context.version}.xlsx")
            cross_plot_workbook.save(excel_file)

            # Delete the plot after saving to Excel
            if os.path.exists(output_file):
                os.remove(output_file)

            logger.info(f"{task}: {self.context.logs_path}")
        except Exception as e:
            logger.error(f"Error {task} workflow task: {e}\n{format_exc()}")
            raise ValueError(f"Error {task} workflow task: {e}")
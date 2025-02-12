import openpyxl.worksheet
from tasks.task import Task
from tasks.task_enum import TASKS
from helpers import task_logger
from traceback import format_exc
from services import GunBarrelService, AnalysisService, WellService
import datetime
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image

class CreateGunBarrelExcel(Task):

    def auto_adjust_column_widths(self, worksheet, headers):
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value in headers:
                        header_length = len(cell.value)
                        header_name = cell.value
                        continue
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            if max_length < header_length:
                adjusted_width = 12
            elif 'Parent Well Name' == header_name:
                adjusted_width = 50
            else:
                adjusted_width = 25
            worksheet.column_dimensions[column].width = adjusted_width

    def execute(self):
        task = TASKS.CREATE_GUN_BARREL_EXCEL.value
        logger = task_logger(task, self.context.logs_path)
        try:
            gun_barrel_service = GunBarrelService(db_path=self.context.db_path)
            analysis_service = AnalysisService(db_path=self.context.db_path)
            well_service = WellService(db_path=self.context.db_path)

            # ####### Create Parent Well Summary Workbook
            gun_barrel_workbook = Workbook()

            # Create Child Well Summary Worksheet
            child_well_summary_worksheet = gun_barrel_workbook.active
            child_well_summary_worksheet.title = "Child Well Summary"

            headers = ['Target Well Ref No.',
                        'Target Well', 
                        'Parent Well Ref No.',
                        'Parent Well API',
                        'Parent Well Name', 
                        'First Production Date', 
                        'Months to First Production', 
                        'Perforated Interval', 
                        'Cumulative Oil', 
                        'Cumulative Oil/Ft', 
                        'Overlap %', 
                        'Overlap CumOil Parent bbls/ft']
            child_well_summary_worksheet.append(headers)

            target_well_set = set()
            for target_well in gun_barrel_service.select_all():
                target_well_set.add(target_well.target_well_api)

            for target_well_api in target_well_set:
                target_well_analysis = analysis_service.get_by_api(target_well_api)
                for gun_barrel_offset_well in gun_barrel_service.select_by_target_well_api(target_well_api):
                    offset_well_analysis = analysis_service.get_by_api(gun_barrel_offset_well.offset_well_api)
                    if offset_well_analysis.gun_barrel_index is None:
                        continue
                    well = well_service.get_by_api(gun_barrel_offset_well.offset_well_api)
                    child_well_summary_worksheet.append([
                            target_well_analysis.gun_barrel_index,
                            target_well_analysis.name,
                            offset_well_analysis.gun_barrel_index,
                            offset_well_analysis.api,
                            offset_well_analysis.name,
                            offset_well_analysis.first_production_date,
                            gun_barrel_offset_well.months_from_first_production,
                            well.perf_interval,
                            f"{well.cumlative_oil} as of {datetime.datetime.now().strftime('%m/%d/%Y')}",
                            gun_barrel_offset_well.cumulative_oil_per_ft,
                            gun_barrel_offset_well.overlap_percentage,
                            gun_barrel_offset_well.overlap_cumulative_oil_ft
                    ])
                child_well_summary_worksheet.append([' ',' ',' ',' ',
                                                     'Total Overlap CumOil bbl/ft and Weighted Avg 3-D',
                                                     ' ',' ',' ',' ',' ',
                                                     0,0
                                                     ])
                child_well_summary_worksheet.append([])
            
            # Adjust the column widths
            self.auto_adjust_column_widths(child_well_summary_worksheet, headers)

            # Center align all cellschild_well_summary_worksheet
            for row in child_well_summary_worksheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(left=Side(border_style='thin', color='FF000000'),
                                         right=Side(border_style='thin', color='FF000000'),
                                         top=Side(border_style='thin', color='FF000000'),
                                         bottom=Side(border_style='thin', color='FF000000'))

            # Set header backgroup color to light blue
            for cell in child_well_summary_worksheet[1]:
                cell.fill = PatternFill(start_color="DDEBF7", end_color="00B0F0", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Insert plot image
            plot_worksheet = gun_barrel_workbook.create_sheet("Plot")
            image_path = os.path.join(self.context.project_path, f"{self.context.project}-gun-barrel-plot-{self.context.version}.png")
            image = Image(image_path)
            image.width = image.width * 0.35 # Scale the width to 50%
            image.height = image.height * 0.35  # Scale the height to 50%

            plot_worksheet.add_image(image, f"A{plot_worksheet.max_row + 0}")
            
            # Save to excel
            excel_file = os.path.join(self.context.project_path, f"{self.context.project}-gun-barrel-plot-{self.context.version}.xlsx")
            gun_barrel_workbook.save(excel_file)

            logger.info(f"{task}: {self.context.logs_path}")    
        except Exception as e:
            logger.error(f"Error {task} workflow task: {e}\n{format_exc()}")
            raise ValueError(f"Error {task} workflow task: {e}")
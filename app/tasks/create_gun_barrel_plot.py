from tasks.task import Task
from tasks.task_enum import TASKS
from helpers import (task_logger, 
                     HandlerWithText, 
                     calculate_angle, 
                     marker_colors,
                     latlon_to_utm_feet)

from services import (AnalysisService, 
                      TargetWellInformationService, 
                      WellService, 
                      GunBarrelService, 
                      GunBarrelTriangleDistancesService, 
                      XYZDistanceService,
                      TexasLandSurveySystemService,
                      NewMexicoLandSurveySystemService)

from traceback import format_exc
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image

BACKGROUND_COLOR = "white"

class CreateGunBarrelPlot(Task):
        
    def auto_adjust_column_widths(self, worksheet, headers):
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value in headers:
                        header_length = len(cell.value)
                        header_name = cell.value
                        continue
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            if max_length < header_length:
                adjusted_width = 12
            elif 'Parent Well Name' == header_name:
                adjusted_width = 45
            elif max_length > 25:
                adjusted_width = max_length
            else:
                adjusted_width = 25
            worksheet.column_dimensions[column].width = adjusted_width

    def enrich_worksheet(self, worksheet, headers):
            # Adjust the column widths
            self.auto_adjust_column_widths(worksheet=worksheet, headers=headers)

            # Center align all cellschild_well_summary_worksheet
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(left=Side(border_style='thin', color='FF000000'),
                                         right=Side(border_style='thin', color='FF000000'),
                                         top=Side(border_style='thin', color='FF000000'),
                                         bottom=Side(border_style='thin', color='FF000000'))

            # Set header backgroup color to light blue
            for cell in worksheet[1]:
                cell.fill = PatternFill(start_color="DDEBF7", end_color="00B0F0", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def execute(self):
        task = TASKS.CREATE_GUN_BARREL_PLOT.value
        logger = task_logger(task, self.context.logs_path)
        try:

            fontsize = 8

            target_well_information_service = TargetWellInformationService(self.context.db_path)
            analysis_service = AnalysisService(self.context.db_path)
            gun_barrel_service = GunBarrelService(self.context.db_path)
            well_service = WellService(self.context.db_path)
            gun_barrel_triangle_distances_service = GunBarrelTriangleDistancesService(self.context.db_path)
            xyz_distance_service = XYZDistanceService(self.context.db_path)
            texas_land_survey_system_service = TexasLandSurveySystemService(self.context._texas_land_survey_system_database_path)
            new_mexico_land_survey_system_service = NewMexicoLandSurveySystemService(self.context._new_mexico_land_survey_system_database_path) 

            target_well = target_well_information_service.get_first_row()

            # Create the plot
            fig, ax = plt.subplots(figsize=(15, 8), facecolor=BACKGROUND_COLOR)

            ax.set_facecolor(BACKGROUND_COLOR)
            ax.grid(True, which='both', color=BACKGROUND_COLOR, linestyle='--', linewidth=0.5)

            # Set plot title
            title = f"{self.context.project.capitalize()} Barrel Plot ({self.context.version})"
            plt.title(title, fontsize=12, fontweight='bold')

            # Set plot demension limits
            shallowest = target_well_information_service.get_shallowest() - self.context.depth_distance_threshold
            deepest = target_well_information_service.get_deepest() + self.context.depth_distance_threshold
            y_min = deepest * -1
            y_max = shallowest * -1
            x_min = -2640
            x_max = 7920
            ax.set_xlim(x_min, x_max)
            ax.set_ylim(y_min, y_max)

            # Draw grid lines
            # plt.yticks(np.arange(y_min, y_max + 1, 500))
            depths = np.arange(y_min, y_max, 100)
            for depth in depths:
                ax.axhline(depth, color='lightgrey', linewidth=0.5, linestyle='--', alpha=0.90, zorder=0)

            plt.xticks(np.arange(x_min, x_max + 1, 2640))
            widths = np.arange(x_min, x_max, 100)
            for width in widths:
                ax.axvline(width, color='lightgrey', linewidth=0.5, linestyle='--', alpha=0.90, zorder=0)  

            # Set axis labels
            if target_well.state == "TX":
                x_axis_label = f"Bottom hole spacing from west line {target_well.state}/{target_well.county}/{target_well.tx_abstract_southwest_corner}) (100 ft intervals)"
                plss = texas_land_survey_system_service.get_by_county_abstract(target_well.county, target_well.tx_abstract_southwest_corner)           
                if plss:
                    x_axis_label = f"Bottom hole spacing from west line {target_well.state}/{target_well.county}/{target_well.tx_abstract_southwest_corner}/{plss.block}/{int(float(plss.section))}) (100 ft intervals)"
                    fnl_grid_x, fnl_grid_y = latlon_to_utm_feet(plss.northeast_latitude, plss.northeast_longitude)
        
            elif target_well.state == "NM":
                x_axis_label = f"Bottom hole spacing from west line {target_well.state}/{target_well.county}/{target_well.nw_township_southwest_corner}/{target_well.nm_range_southwest_corner}/{int(float(target_well.nm_tx_section_southwest_corner))}) (100 ft intervals)"
                township = int(target_well.nw_township_southwest_corner[:-1])
                township_direction = target_well.nw_township_southwest_corner[-1]
                nm_range = int(target_well.nm_range_southwest_corner[:-1])
                range_direction = target_well.nm_range_southwest_corner[-1]
                section = int(target_well.nm_tx_section_southwest_corner)
                plss = new_mexico_land_survey_system_service.get_by_township_range_section(township=township, township_direction=township_direction, range=nm_range, range_direction=range_direction, section=section)            
                if plss:
                    fnl_grid_x, fnl_grid_y = latlon_to_utm_feet(plss.northeast_latitude, plss.northeast_longitude)

            ax.set_xlabel(x_axis_label)
            ax.set_ylabel(f"Depth below mean sea level (100 ft intervals)")

            # Plot west section line
            specific_x = 0
            ax.axvline(specific_x, color='blue', linewidth=0.5, linestyle='--', alpha=0.75)
            label_y_position = y_min
            ax.text(specific_x, label_y_position, x_axis_label, color='black', fontsize=fontsize, ha='center', va='bottom')

            legend_elements = []  
            saved_i = 1
            wells = []
            api_to_index = {}
            text = ""

            # Plot target wells
            target_wells = target_well_information_service.get_all()
            color = "orange"
            text_color = "black"
            for target_well in target_wells:
                target_well_analysis = analysis_service.get_by_name(target_well.name)
                wells.append(target_well_analysis)
                api_to_index[target_well_analysis.api] = target_well.id
                saved_i = target_well.id
                target_well_analysis.gun_barrel_index = saved_i
                analysis_service.update(target_well_analysis)
                ax.scatter(target_well_analysis.gun_barrel_x, target_well_analysis.subsurface_depth, facecolor=color, s=100, edgecolor=BACKGROUND_COLOR, linewidth=0.5, alpha=0.70, zorder=3)
                ax.annotate(str(saved_i), (target_well_analysis.gun_barrel_x, target_well_analysis.subsurface_depth), color=text_color, fontsize=fontsize, ha='center', va='center', weight='bold', zorder=6)
                y_offset = -100
                # ax.annotate(str(target_well_analysis.first_production_date), (target_well_analysis.gun_barrel_x, target_well_analysis.subsurface_depth + y_offset), color='black', fontsize=fontsize, ha='center', va='center', alpha=0.50, zorder=6)
                # Annotate the perforation interval
                y_offset = 100
                interval = str(target_well.afe_landing_zone) if target_well.afe_landing_zone else ""
                ax.annotate(f"{interval}", (target_well_analysis.gun_barrel_x, target_well_analysis.subsurface_depth + y_offset), color='black', fontsize=fontsize, ha='center', va='center', alpha=0.50, zorder=6)
                # Add legend elements
                legend_elements.append(Line2D([0], [0], marker='o', color='w', markerfacecolor=color, markeredgecolor=BACKGROUND_COLOR, markersize=10, linestyle='None', alpha=0.70, zorder=5))

            # Plot target well's adjacet offset wells
            for target_well in target_wells:
                target_well_analysis = analysis_service.get_by_name(target_well.name)
                if target_well_analysis.adjacent_1 is not None:
                    adjacent_1_analysis = analysis_service.get_by_name(target_well_analysis.adjacent_1)
                    if adjacent_1_analysis.api in api_to_index:
                        continue
                    saved_i = saved_i + 1
                    adjacent_1_analysis.gun_barrel_index = saved_i
                    analysis_service.update(adjacent_1_analysis)
                    wells.append(adjacent_1_analysis)
                    api_to_index[adjacent_1_analysis.api] = saved_i
                    well = well_service.get_by_api(adjacent_1_analysis.api)
                    color, text_color = marker_colors(well.status)
                    ax.scatter(adjacent_1_analysis.gun_barrel_x, adjacent_1_analysis.subsurface_depth, facecolor=color, s=100, edgecolor=BACKGROUND_COLOR, linewidth=0.5, alpha=0.70, zorder=3)
                    ax.annotate(str(saved_i), (adjacent_1_analysis.gun_barrel_x, adjacent_1_analysis.subsurface_depth), color=text_color, fontsize=fontsize, ha='center', va='center', weight='bold', zorder=6)
                    y_offset = -100
                    # ax.annotate(str(adjacent_1_analysis.first_production_date), (adjacent_1_analysis.gun_barrel_x, adjacent_1_analysis.subsurface_depth + y_offset), color='black', fontsize=fontsize, ha='center', va='center', alpha=0.50, zorder=6)
                    # Annotate the perforation interval
                    y_offset = 100
                    interval = str(adjacent_1_analysis.interval) if adjacent_1_analysis.interval else str(well.interval)
                    ax.annotate(f"{interval}", (adjacent_1_analysis.gun_barrel_x, adjacent_1_analysis.subsurface_depth + y_offset), color='black', fontsize=fontsize, ha='center', va='center', alpha=0.50, zorder=6)
                    # Add legend elements
                    legend_elements.append(Line2D([0], [0], marker='o', color='w', markerfacecolor=color, markeredgecolor=BACKGROUND_COLOR, markersize=10, linestyle='None', alpha=0.70, zorder=5))
                if target_well_analysis.adjacent_2 is not None:
                    adjacent_2_analysis = analysis_service.get_by_name(target_well_analysis.adjacent_2)
                    if adjacent_2_analysis.api in api_to_index:
                        continue
                    saved_i = saved_i + 1
                    adjacent_2_analysis.gun_barrel_index = saved_i
                    analysis_service.update(adjacent_2_analysis)
                    wells.append(adjacent_2_analysis)
                    api_to_index[adjacent_2_analysis.api] = saved_i
                    well = well_service.get_by_api(adjacent_2_analysis.api)
                    color, text_color = marker_colors(well.status)
                    ax.scatter(adjacent_2_analysis.gun_barrel_x, adjacent_2_analysis.subsurface_depth, facecolor=color, s=100, edgecolor=BACKGROUND_COLOR, linewidth=0.5, alpha=0.70, zorder=3)
                    ax.annotate(str(saved_i), (adjacent_2_analysis.gun_barrel_x, adjacent_2_analysis.subsurface_depth), color=text_color, fontsize=fontsize, ha='center', va='center', weight='bold', zorder=6)
                    y_offset = -100
                    # ax.annotate(str(adjacent_2_analysis.first_production_date), (adjacent_2_analysis.gun_barrel_x, adjacent_2_analysis.subsurface_depth + y_offset), color='black', fontsize=fontsize, ha='center', va='center', alpha=0.50, zorder=6)
                    # Annotate the perforation interval
                    y_offset = 100
                    interval = str(adjacent_2_analysis.interval) if adjacent_2_analysis.interval else str(well.interval)
                    ax.annotate(f"{interval}", (adjacent_2_analysis.gun_barrel_x, adjacent_2_analysis.subsurface_depth + y_offset), color='black', fontsize=fontsize, ha='center', va='center', alpha=0.50, zorder=6)
                    # Add legend elements
                    legend_elements.append(Line2D([0], [0], marker='o', color='w', markerfacecolor=color, markeredgecolor=BACKGROUND_COLOR, markersize=10, linestyle='None', alpha=0.70, zorder=5))

            # Draw triangle distances between target well and adjacent offset wells
            for target_well in target_wells:
                target_well_analysis = analysis_service.get_by_name(target_well.name)
                if target_well_analysis.adjacent_1 is not None:
                    adjacent_1_analysis = analysis_service.get_by_name(target_well_analysis.adjacent_1)
                    x_distance = abs(target_well_analysis.gun_barrel_x - adjacent_1_analysis.gun_barrel_x)
                    y_distance = abs(target_well_analysis.subsurface_depth - adjacent_1_analysis.subsurface_depth)
                    hypotenuse_distance = int(np.sqrt(x_distance**2 + y_distance**2))
                    if hypotenuse_distance < 1800:
                        # Calculate midpoints for hypotenuse
                        mid_x = (target_well_analysis.gun_barrel_x + adjacent_1_analysis.gun_barrel_x) / 2
                        mid_y = (target_well_analysis.subsurface_depth + adjacent_1_analysis.subsurface_depth) / 2
                        ax.plot([target_well_analysis.gun_barrel_x, adjacent_1_analysis.gun_barrel_x],
                                [target_well_analysis.subsurface_depth, adjacent_1_analysis.subsurface_depth],
                                color='black', linestyle='--', linewidth=0.5)
                        # Add hypotenuse label
                        ax.text(mid_x, mid_y, f"{hypotenuse_distance}", fontsize=fontsize, ha='right', va='bottom', alpha=0.90, zorder=8)
                if target_well_analysis.adjacent_2 is not None:
                    adjacent_2_analysis = analysis_service.get_by_name(target_well_analysis.adjacent_2)
                    x_distance = abs(target_well_analysis.gun_barrel_x - adjacent_2_analysis.gun_barrel_x)
                    y_distance = abs(target_well_analysis.subsurface_depth - adjacent_2_analysis.subsurface_depth)
                    hypotenuse_distance = int(np.sqrt(x_distance**2 + y_distance**2))
                    if hypotenuse_distance < 1800:
                        # Calculate midpoints for hypotenuse
                        mid_x = (target_well_analysis.gun_barrel_x + adjacent_2_analysis.gun_barrel_x) / 2
                        mid_y = (target_well_analysis.subsurface_depth + adjacent_2_analysis.subsurface_depth) / 2
                        ax.plot([target_well_analysis.gun_barrel_x, adjacent_2_analysis.gun_barrel_x],
                            [target_well_analysis.subsurface_depth, adjacent_2_analysis.subsurface_depth],
                            color='black', linestyle='--', linewidth=0.5)
                        ax.text(mid_x, mid_y, f"{hypotenuse_distance}", fontsize=fontsize, ha='right', va='bottom', alpha=0.90, zorder=8)

            # Plot of balane of the offset wells
            gun_barrel_distances = gun_barrel_triangle_distances_service.select_all()
            for gun_barrel_distance in gun_barrel_distances:
                if gun_barrel_distance.hypotenuse > int(self.context.hypotenuse_distance_threshold):
                    continue
                offset_well_analysis = analysis_service.get_by_api(gun_barrel_distance.offset_well_api)
                if offset_well_analysis.api in api_to_index:
                    continue
                if offset_well_analysis.gun_barrel_x < x_min or offset_well_analysis.gun_barrel_x > x_max:
                    continue
                saved_i = saved_i + 1
                offset_well_analysis.gun_barrel_index = saved_i
                analysis_service.update(offset_well_analysis)
                wells.append(offset_well_analysis)
                api_to_index[offset_well_analysis.api] = saved_i
                well = well_service.get_by_api(offset_well_analysis.api)
                color, text_color = marker_colors(well.status)
                ax.scatter(offset_well_analysis.gun_barrel_x, offset_well_analysis.subsurface_depth, facecolor=color, s=100, edgecolor=BACKGROUND_COLOR, linewidth=0.5, alpha=0.70, zorder=3)
                ax.annotate(str(saved_i), (offset_well_analysis.gun_barrel_x, offset_well_analysis.subsurface_depth), color=text_color, fontsize=fontsize, ha='center', va='center', weight='bold', zorder=6)
                y_offset = -100
                # ax.annotate(str(offset_well_analysis.first_production_date), (offset_well_analysis.gun_barrel_x, offset_well_analysis.subsurface_depth + y_offset), color='black', fontsize=fontsize, ha='center', va='center', alpha=0.50, zorder=6)
                # Annotate the perforation interval
                y_offset = 100
                interval = str(offset_well_analysis.interval) if offset_well_analysis.interval else str(well.interval)
                ax.annotate(f"{interval}", (offset_well_analysis.gun_barrel_x, offset_well_analysis.subsurface_depth + y_offset), color='black', fontsize=fontsize, ha='center', va='center', alpha=0.50, zorder=6)
                # Add legend elements
                legend_elements.append(Line2D([0], [0], marker='o', color='w', markerfacecolor=color, markeredgecolor=BACKGROUND_COLOR, markersize=10, linestyle='None', alpha=0.70, zorder=5))
                
            # Try to center the plot on the target wells
            smallest_positive_x = 5000
            for well in wells:
                if well.gun_barrel_x > 0 and well.gun_barrel_x < smallest_positive_x:
                    smallest_positive_x = well.gun_barrel_x  

            largest_negative_x = 0
            for well in wells:
                if well.gun_barrel_x < 0 and well.gun_barrel_x < largest_negative_x:
                    largest_negative_x = well.gun_barrel_x  

            # x_min = x_min + (smallest_positive_x + largest_negative_x)
            # x_max = x_max + (smallest_positive_x + largest_negative_x)
            # ax.set_xlim(x_min, x_max)

            # Add custom legend
            handlers = {legend_elements[i-1]: HandlerWithText(str(i)) for i in range(1, len(legend_elements) + 1)}
            custom_legend = ax.legend(
                handles=legend_elements,
                labels=[f"{well.name}" for well in wells],
                loc='upper center',
                fontsize=fontsize,
                bbox_to_anchor=(0.5, -0.14),
                handler_map=handlers,
                borderaxespad=0.1,
                ncol=(len(wells)))

            custom_legend.get_frame().set_facecolor(BACKGROUND_COLOR)
            ax.add_artist(custom_legend)

            # Add custom annotation
            text =  f"Landing Zone (LZ) are shown at the top of markers.\n\n"
            text += f"Targets:orange, Proudcing:green\n\n"
            text += f"From-To (FPD, LZ, Perf Int, horizontal, vertical, 3-D, \n"
            text += f"months from FPD, cumoil (MBO), cumoil bbls/ft, overlap%, \n"
            text += f"cumoil*overlap)\n\n"

            pairs = []
            for target_well in target_wells:
                target_well_analysis = analysis_service.get_by_name(target_well.name)
                gun_barrel_distances = gun_barrel_triangle_distances_service.select_by_target_api(target_well_api=target_well_analysis.api)
                for gun_barrel_distance in gun_barrel_distances:
                    if (gun_barrel_distance.target_well_api in api_to_index and 
                        gun_barrel_distance.offset_well_api in api_to_index and
                        gun_barrel_distance.hypotenuse < int(self.context.hypotenuse_distance_threshold)):
                        pair = (gun_barrel_distance.target_well_api, gun_barrel_distance.offset_well_api)
                        pairs.append(pair)
                        gun_barrel = gun_barrel_service.select_by_target_offset_well_api(gun_barrel_distance.target_well_api, gun_barrel_distance.offset_well_api)
                        offset_well = well_service.get_by_api(gun_barrel_distance.offset_well_api)
                        offset_well_analysis = analysis_service.get_by_api(gun_barrel_distance.offset_well_api)
                        text += f"{api_to_index[gun_barrel_distance.target_well_api]}-{api_to_index[gun_barrel_distance.offset_well_api]} "
                        text += f"("
                        text += f"{offset_well.first_production_date}, "
                        text += f"{offset_well_analysis.interval}, "
                        text += f"{offset_well.perf_interval}, "
                        text += f"{gun_barrel_distance.adjacent}, "
                        text += f"{gun_barrel_distance.opposite}, "
                        text += f"{gun_barrel_distance.hypotenuse}, "
                        text += f"{gun_barrel.months_from_first_production}, "
                        text += f"{int(offset_well.cumlative_oil/1000)}, "
                        text += f"{gun_barrel.cumulative_oil_per_ft}, "
                        text += f"{gun_barrel.overlap_percentage}, "
                        text += f"{gun_barrel.overlap_cumulative_oil_ft}"
                        text += f")\n"

            plt.text(plt.gca().get_xlim()[1] + 0.015 * (plt.gca().get_xlim()[1] - plt.gca().get_xlim()[0]),
                     plt.gca().get_ylim()[0] - 0.015 * (plt.gca().get_ylim()[1] - plt.gca().get_ylim()[0]),  
                     text,
                     fontsize=10, 
                     color='black')
            
            # Save the plot
            output_file = os.path.join(self.context.project_path, f"{self.context.project}-gun-barrel-plot-{self.context.version}.png")
            plt.savefig(output_file, dpi=300, bbox_inches='tight', facecolor=BACKGROUND_COLOR)
            plt.close(fig)

            # Create the Excel Spreadsheet

            # ####### Create Parent Well Summary Workbook
            gun_barrel_workbook = Workbook()

            # Create Well Table Worksheet
            well_table_worksheet = gun_barrel_workbook.active
            well_table_worksheet.title = "Well Table"
            headers = ['Ref No.',
                        'API',
                        'Well Name', 
                        'EnvOperator', 
                        'Prism_EnvInterval', 
                        'First Production Date', 
                        'TVD_FT', 
                        'PerfInterval_FT',
                        'LateralLength_FT',
                        'ProppantIntensity_LBSPerFT',
                        'Type',
                        'Group by Radius']
            
            well_table_worksheet.append(headers)

            for key, value in api_to_index.items():
                if '11-111' in key:
                    continue
                well = well_service.get_by_api(key)
                well_table_worksheet.append([
                    value,
                    well.api,
                    well.name,
                    well.operator,
                    well.interval,
                    well.first_production_date,
                    well.total_vertical_depth,
                    well.perf_interval,
                    well.lateral_length,
                    well.proppant_intensity,
                    '',
                    ''
                ])
            
            self.enrich_worksheet(well_table_worksheet, headers)

            # Create Plot Data Worksheet
            plot_data_worksheet = gun_barrel_workbook.create_sheet("Plot Data")
            headers = ['Well Ref Number',	
                       'API',	
                       'Target Well',	
                       'TVD Subsea (SS)',	
                       'FWL Section',	
                       'FNL Section']
            plot_data_worksheet.append(headers)
            for key, value in api_to_index.items():
                target_well_analysis = analysis_service.get_by_api(key)
                xyz_distance = xyz_distance_service.get_by_reference_target_well(reference_well_api='00-000-00000', target_well_api=key)
                fwl = 'TBD'
                fnl = 'TBD'
                if xyz_distance:
                    fwl = str(xyz_distance.end_x)
                if fnl_grid_y:
                    fnl = str(int(fnl_grid_y - target_well_analysis.lateral_end_grid_y))
                if '11-111' in key:
                    key = ''
                plot_data_worksheet.append([
                    value,
                    key,
                    target_well_analysis.name,    
                    target_well_analysis.subsurface_depth,
                    fwl,
                    fnl
                ])

            self.enrich_worksheet(plot_data_worksheet, headers)

            # Create Calculated Data
            calculated_data_worksheet = gun_barrel_workbook.create_sheet("Calculated Data") 
            header = [
                'target_well_index',
                'offset_well_index',
                'Target Well Name',	
                'Offset Well Name',	
                'Targete Well First Prod Date',	
                'Offset Well First Prod Date',	
                'Months to First Production',	
                'Target Well Perf Interval',	
                'Offset Well Perf Interval',	
                'Offset Well Cum Oil Date',	
                'overlap_feet',	
                'overlap_percentage',
                'Cum Oil',	
                'cumulative_oil_per_ft',	
                'overlap_cumulative_oil_ft',	
                'months_from_first_production',	
                'horizontal_distance',	
                'vertical_distance',	
                'three_d_distance']
            calculated_data_worksheet.append(header)

            for target_well in target_wells:
                target_well_analysis = analysis_service.get_by_name(target_well.name)
                sum_of_overlap_cumulative_oil_ft: int = 0
                sum_3d_distance: int = 0
                count = 0
                for pair in pairs:
                    if target_well_analysis.api == pair[0]:
                        offset_well_api = pair[1]
                        offset_well_analysis = analysis_service.get_by_api(offset_well_api)
                        gun_barrel = gun_barrel_service.select_by_target_offset_well_api(target_well_analysis.api, offset_well_api)
                        gun_barrel_distance = gun_barrel_triangle_distances_service.select_by_target_offset_api(target_well_analysis.api, offset_well_api)
                        well = well_service.get_by_api(offset_well_api)
                        target_well = target_well_information_service.get_by_name(target_well_analysis.name)
                        calculated_data_worksheet.append([
                                target_well_analysis.gun_barrel_index,
                                offset_well_analysis.gun_barrel_index,
                                target_well_analysis.name,
                                offset_well_analysis.name,
                                target_well_analysis.first_production_date,
                                offset_well_analysis.first_production_date,
                                gun_barrel.months_from_first_production,
                                target_well.surveys_preforated_interval_ft,
                                well.perf_interval,
                                "TBD",
                                gun_barrel.overlap_feet,
                                gun_barrel.overlap_percentage,
                                well.cumlative_oil,
                                gun_barrel.cumulative_oil_per_ft,
                                gun_barrel.overlap_cumulative_oil_ft,
                                gun_barrel.months_from_first_production,
                                gun_barrel_distance.adjacent,
                                gun_barrel_distance.opposite,
                                gun_barrel_distance.hypotenuse
                            ])
                        sum_of_overlap_cumulative_oil_ft += gun_barrel.overlap_cumulative_oil_ft
                        sum_3d_distance += gun_barrel_distance.hypotenuse
                        count += 1

            self.enrich_worksheet(calculated_data_worksheet, header)

            # Insert plot image
            plot_worksheet = gun_barrel_workbook.create_sheet("Plot")
            image_path = os.path.join(self.context.project_path, f"{self.context.project}-gun-barrel-plot-{self.context.version}.png")
            image = Image(image_path)
            image.width = image.width * 0.35 # Scale the width to 50%
            image.height = image.height * 0.35  # Scale the height to 50%

            plot_worksheet.add_image(image, f"A{plot_worksheet.max_row + 0}")

            # Save to excel
            excel_file = os.path.join(self.context.project_path, f"{self.context.project}-gun-barrel-plot-{self.context.version}.xlsx")
            gun_barrel_workbook.save(excel_file)

            # Delete the plot after saving to Excel
            if os.path.exists(output_file):
                os.remove(output_file)

            logger.info(f"{task}: {self.context.logs_path}")
        except Exception as e:
            message = f"Error {task} workflow task: {e}"
            logger.error(msg=message)
            raise ValueError(e)
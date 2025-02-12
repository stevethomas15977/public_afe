import base64
import requests
import os
import cv2
import numpy as np
from PIL import Image, ImageEnhance
from dotenv import load_dotenv
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

load_dotenv()

# Load OpenAI API Key
api_key = os.getenv("OPENAI_API_KEY")


def encode_image(image_path):
    """Encodes the image to a base64 string."""
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode('utf-8')


def preprocess_image(image_path):
    """Preprocesses the image to enhance its quality."""
    image = cv2.imread(image_path)
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    filtered = cv2.bilateralFilter(gray, 11, 17, 17)
    pil_image = Image.fromarray(filtered)
    enhancer = ImageEnhance.Contrast(pil_image)
    enhanced_image = enhancer.enhance(2.0)
    processed_image = np.array(enhanced_image)
    temp_path = "processed_image.png"
    cv2.imwrite(temp_path, processed_image)
    return temp_path


def get_image_text(image_path, temperature=0, top_p=0.1, max_tokens=4096):
    """Gets the text content from the image using the OpenAI API."""
    preprocessed_image_path = preprocess_image(image_path)
    base64_image = encode_image(preprocessed_image_path)
    
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}"
    }
    
    payload = {
        "model": "gpt-4o-mini",
        "messages": [
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": "You are a helpful assistant. Parse the provided image and return the complete text content. If the response is too long, continue until all text is processed."
                    },
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/jpeg;base64,{base64_image}"
                        }
                    }
                ]
            }
        ],
        "max_tokens": max_tokens,
        "temperature": temperature,
        "top_p": top_p
    }

    response = requests.post("https://api.openai.com/v1/chat/completions", headers=headers, json=payload)
    response_data = response.json()
    content = response_data['choices'][0]['message']['content']

    while response_data['choices'][0]['finish_reason'] == 'length':
        prompt_continuation = f"Continue from: {content[-100:]}"
        payload_continuation = {
            "model": "gpt-4o-mini",
            "messages": [
                {
                    "role": "user",
                    "content": prompt_continuation
                }
            ],
            "max_tokens": max_tokens,
            "temperature": temperature,
            "top_p": top_p
        }
        response = requests.post("https://api.openai.com/v1/chat/completions", headers=headers, json=payload_continuation)
        response_data = response.json()
        content += response_data['choices'][0]['message']['content']

    return content

def extract_text_between_backticks(text):
    """Extracts text between triple backticks."""
    pattern = r'```(.*?)```'
    match = re.search(pattern, text, re.DOTALL)
    if match:
        return match.group(1).strip()
    else:
        return None

def parse_sections(extracted_text):
    """Parses the text into sections based on predefined headers."""
    sections = ["SURFACE LOCATION", "PENETRATION POINT", "FIRST TAKE POINT", "LAST TAKE POINT", "END OF TERMINUS"]
    pattern = r'(' + '|'.join(map(re.escape, sections)) + r')'
    split_text = re.split(pattern, extracted_text)
    
    parsed_sections = {}
    for i in range(1, len(split_text), 2):
        header = split_text[i]
        content = split_text[i + 1].strip()
        parsed_sections[header] = content
    
    return parsed_sections

def parse_section(text) -> list[str]: 
    lines = text.strip().split('\n')
    return [line.strip() for line in lines if line.strip()]

def extract_attributes(section_text):
    """Extracts attributes from a section of text using predefined regex patterns."""
    attribute_regex_patterns = {
        "WELL NO": r"WELL NO\. (.*)$",
        "FNL": r"(\d{3,4})' FNL",
        "FSL": r"(\d{3,4})' FSL",
        "FEL": r"(\d{3,4})' FEL",
        "FWL": r"(\d{3,4})' FWL",
        "Abstract": r"A-[^,]+",
        "Section": r"Section ([^.,]+)[.,]",
        "Block": r"Blk\. ([^,]+),",
        "Township": r"TSP [^,]*",
        "Survey": r", ([^,]*?) SURVEY,",
        "County": r"SURVEY, (.*?) Co\.",
        "X": r"X= ([^']*)\'",
        "Y": r"Y= ([^']*)\'",
        "Latitude1": r"Lat\.\= (.*?) Long",
        "Longitude1": r"Long\.\= (.*)$",
        "Latitude2": r"Lat\.\: (.*?) Long",
        "Longitude2": r"Long\.\: (.*)$",
        "Ground Elevation": r"Ground Elevation: ([^']*)\'",
    }
    
    section_data = {}
    for attribute, regex in attribute_regex_patterns.items():
        for line in parse_section(section_text):
            match = re.search(regex, line)
            if match:
                if match.groups():
                    if ((attribute == "Latitude1" and "\"" in match.group(1)) or 
                        (attribute == "Longitude1" and "\"" in match.group(1)) or
                        (attribute == "Latitude2" and "\"" in match.group(1)) or 
                        (attribute == "Longitude2" and "\"" in match.group(1))):
                        section_data[attribute] = dms_to_decimal(match.group(1))
                    else:
                        section_data[attribute] = match.group(1)
                else:
                    section_data[attribute] = match.group(0)
                break
    
    return section_data

def dms_to_decimal(dms):
    pattern = re.compile(r'([-+]?\d+)°(\d+)[\'|′](\d+\.\d+)"')
    match = pattern.match(dms)
    if not match:
        raise ValueError("Invalid DMS format")
    degrees, minutes, seconds = match.groups()
    degrees = float(degrees)
    minutes = float(minutes)
    seconds = float(seconds)
    decimal_degrees = abs(degrees) + (minutes / 60) + (seconds / 3600)
    if degrees < 0:
        decimal_degrees = -decimal_degrees
    return decimal_degrees

def save_to_excel(parsed_data, excel_path):
    """Saves the parsed data into an Excel file with each section as a separate worksheet and formats the columns."""
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        for section, attributes in parsed_data.items():
            df = pd.DataFrame([attributes])
            df.to_excel(writer, sheet_name=section, index=False)

    # Open the workbook to adjust formatting
    workbook = load_workbook(excel_path)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # Auto-adjust column width
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Center align all cells
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    workbook.save(excel_path)

def main(image_dir, excel_dir):
    for filename in os.listdir(image_dir):
        if filename.endswith(".png"):
            image_file = os.path.splitext(filename)[0]
            image_path = os.path.join(image_dir, filename)
            excel_path = os.path.join(excel_dir, f"{image_file}.xlsx")

            content = get_image_text(image_path)
            print(content)
            extracted_text = extract_text_between_backticks(content)
            
            if extracted_text:
                parsed_sections = parse_sections(extracted_text)
                sections_dict = {}
                for section, text in parsed_sections.items():
                    parsed_data = extract_attributes(text)
                    parsed_data["text"] = text
                    sections_dict[section] = parsed_data
                save_to_excel(sections_dict, excel_path)
            else:
                print(f"No text found between backticks for {filename}.")


if __name__ == "__main__":
    image_dir = "/home/ubuntu/afe/projects/moosehorn-3-mile/target_well_information"
    excel_dir = "/home/ubuntu/afe/projects/moosehorn-3-mile/target_well_information"
    main(image_dir, excel_dir)
